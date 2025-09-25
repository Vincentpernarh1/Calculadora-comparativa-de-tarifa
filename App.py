import webview
import json
import os
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import openpyxl
import numpy as np
import sys


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path).replace('\\', '/') # Use forward slashes for HTML


# !------------------------------ API code and conditions -----------------------------------------!
class Api:

    
    def __init__(self):
        self.base_folder = None
        self.fluxo_data = {}
        self.last_results_df = None
        self._window = None
    
   
    def _parse_transporter_name(self, filename):
        """Extracts a cleaned-up transporter name from the filename."""
        match = re.search(r'_(.*?)\.', filename)
        if match:
            return match.group(1).replace('_', ' ').title()
        return os.path.splitext(filename)[0].replace('_', ' ').title()
    
    
    def select_folder(self):
        """Handles the main folder selection using pywebview's native dialog."""
        try:
            if not self._window:
                return {'success': False, 'message': 'Erro: A referência da janela não foi configurada.'}
            
            dialog_result = self._window.create_file_dialog(webview.FileDialog.FOLDER)

            if not dialog_result:
                return {'success': False, 'message': 'Nenhuma pasta foi selecionada.'}
            
            folder_path = dialog_result[0]
            self.base_folder = folder_path
            self.fluxo_data = {}
            
            fluxos = [d for d in os.listdir(self.base_folder) if os.path.isdir(os.path.join(self.base_folder, d))]
            
            if not fluxos:
                return {'success': False, 'message': 'Nenhum fluxo (subpasta) encontrado na pasta selecionada.'}
                
            return {'success': True, 'path': self.base_folder, 'fluxos': sorted(fluxos)}
            
        except Exception as e:
            print(f"PYTHON ERROR in select_folder: {e}") 
            return {'success': False, 'message': f'Ocorreu um erro no Python: {e}'}

    def get_filters_for_fluxo(self, fluxo_name):
        """
        Reads all Excel files in a given fluxo, processes them into a tidy (long) DataFrame,
        and returns the available filters for the UI.
        """



        def normalize_vehicle_name(name):
            """
            Normalizes vehicle names by handling different spellings, cases, and variations,
            while keeping distinct vehicle types separate.
            """
            if not name or not str(name).strip(): return None
            clean_name = str(name).strip().upper()
            if 'BITREM' in clean_name: return 'BITREM'
            if 'VANDERLEIA' in clean_name: return 'VANDERLEIA'
            if 'CARRETA' in clean_name: return 'CARRETA'
            if 'VAN' in clean_name or 'DUCATO' in clean_name: return 'VAN'
            if '3/4' in clean_name or '0.75' in clean_name: return '3/4'
            if 'TOCO' in clean_name: return 'TOCO'
            if 'TRUCK' in clean_name: return 'TRUCK'
            if 'FIORINO' in clean_name: return 'FIORINO'
            return clean_name

        if fluxo_name in self.fluxo_data:
            return self.fluxo_data[fluxo_name]['filters_response']

        fluxo_path = os.path.join(self.base_folder, fluxo_name)
        all_melted_dfs = []




        if '04. MILK RUN' in fluxo_name:
            for file_name in os.listdir(fluxo_path):
                if not file_name.lower().endswith(('.xlsx', '.xls')) or file_name.startswith('~$'):
                    continue
                
                file_path = os.path.join(fluxo_path, file_name)
                try:
                    full_df = pd.read_excel(file_path, header=None, engine='openpyxl', keep_default_na=False, na_values=[''])

                    header_row_idx = -1
                    faixa_col_idx = -1
                    
                    for r_idx in range(min(10, len(full_df))):
                        row_as_str = full_df.iloc[r_idx].astype(str).str.strip().str.upper()
                        matches = row_as_str[row_as_str == 'FAIXA KM']
                        if not matches.empty:
                            header_row_idx = r_idx
                            faixa_col_idx = matches.index[0]
                            break
                    
                    if header_row_idx == -1:
                        print(f"Skipping file {file_name}: Could not find 'FAIXA KM' header.")
                        continue

                    rt_info_row_idx = header_row_idx - 1
                    data_start_row_idx = header_row_idx + 1

                    if rt_info_row_idx < 0 or data_start_row_idx >= len(full_df):
                        print(f"Skipping file {file_name}: Invalid structure around 'FAIXA KM' header.")
                        continue

                    # FIX 1: Add .ffill() to carry over the trip type to all vehicles
                    round_trip_info = full_df.iloc[rt_info_row_idx].ffill()
                    vehicle_headers = full_df.iloc[header_row_idx]
                    data_df = full_df.iloc[data_start_row_idx:]
                    
                    vehicle_col_indices = [
                        idx for idx, val in enumerate(vehicle_headers)
                        if idx > faixa_col_idx and pd.notna(val) and str(val).strip() != ""
                    ]

                    processed_rows = []
                    for _, data_row in data_df.iterrows():
                        faixa_km_str = str(data_row.iloc[faixa_col_idx]).strip()
                        distancia_min = distancia_max = None

                        match = re.search(r'(\d+)\s*(?:a|-|até)\s*(\d+)', faixa_km_str, re.IGNORECASE)
                        if match:
                            distancia_min, distancia_max = int(match.group(1)), int(match.group(2))
                        else:
                            match = re.search(r'acima de\s*(\d+)', faixa_km_str, re.IGNORECASE)
                            if match:
                                distancia_min, distancia_max = int(match.group(1)), float(200)
                            else:
                                match = re.search(r'até\s*(\d+)', faixa_km_str, re.IGNORECASE)
                                if match:
                                    distancia_min, distancia_max = 0, int(match.group(1))

                        if distancia_min is None:
                            continue

                        for col_idx in vehicle_col_indices:
                            try:
                                vehicle = vehicle_headers.iloc[col_idx]
                                viagem = round_trip_info.iloc[col_idx]
                                tarifa = data_row.iloc[col_idx]

                                if pd.notna(tarifa) and str(tarifa).strip() not in ("", "nan") and pd.notna(vehicle) and pd.notna(viagem):
                                    viagem_str = str(viagem).upper().strip()
                                    if 'ROUND' in viagem_str:
                                        viagem_code = 'RT'
                                    elif 'ONE WAY' in viagem_str or 'OW' in viagem_str:
                                        viagem_code = 'OW'
                                    else:
                                        viagem_code = viagem_str
                                    
                                    processed_rows.append({
                                        'Nomeacao': 'N/A', 'Fornecedor': 'N/A', 'Origem': 'N/A',
                                        'LocalColeta': 'N/A', 'Destino': 'N/A',
                                        'DistanciaMin': distancia_min,
                                        'DistanciaMax': distancia_max,
                                        'Transportadora': self._parse_transporter_name(file_name),
                                        'Veiculo': str(vehicle), 
                                        'Viagem': viagem_code,
                                        'Tarifa': tarifa, 'Chave': 'N/A & N/A'
                                    })
                            except Exception as inner_e:
                                print(f"Error processing a vehicle column in {file_name}: {inner_e}")
                                continue
                        
                    if processed_rows:
                        all_melted_dfs.append(pd.DataFrame(processed_rows))

                except Exception as e:
                    print(f"Error processing file {file_path} for 'MILK RUN': {e}")
                    continue

        elif 'FAIXA' in fluxo_name:  # Catch '02. FAIXA', 'FAIXA' etc.
            for file_name in os.listdir(fluxo_path):
                if not file_name.lower().endswith(('.xlsx', '.xls')):
                    continue
                file_path = os.path.join(fluxo_path, file_name)
                try:
                    full_df = pd.read_excel(file_path, header=None, engine='openpyxl', keep_default_na=False, na_values=[''])

                    vehicles = full_df.iloc[0].ffill()
                    headers = full_df.iloc[1]

                    # Start reading from row 4 (index 3)
                    data_df = full_df.iloc[3:].copy()
                    data_df.columns = headers

                    rename_map = {}
                    for col in data_df.columns:
                        col_str = str(col).lower().strip()
                        if col_str == 'origem':
                            rename_map[col] = 'Origem'
                        elif col_str == 'destino':
                            rename_map[col] = 'Destino'
                    data_df.rename(columns=rename_map, inplace=True)

                    origem_indices = [i for i, col in enumerate(data_df.columns) if col == 'Origem']
                    destino_indices = [i for i, col in enumerate(data_df.columns) if col == 'Destino']

                    if not origem_indices or not destino_indices:
                        print(f"Skipping file {file_name} due to missing 'Origem' or 'Destino' columns.")
                        continue

                    origem_idx = origem_indices[0]
                    destino_idx = destino_indices[0]

                    processed_rows = []
                    for _, data_row in data_df.iterrows():
                        origem = data_row.iloc[origem_idx]
                        destino_full = data_row.iloc[destino_idx]

                        destino = destino_full
                        distancia_min = None
                        distancia_max = None

                        if isinstance(destino_full, str):
                            # Extract prefix (UF code like "SP", "MG", etc.)
                            destino = destino_full[:2].strip()

                            # Extract "de X a Y" using regex
                            match = re.search(r'de\s*(\d+)\s*a\s*(\d+)', destino_full, flags=re.IGNORECASE)
                            if match:
                                try:
                                    distancia_min = int(match.group(1))
                                    distancia_max = int(match.group(2))
                                except Exception as e:
                                    print(f"Erro ao converter Distancia em '{destino_full}': {e}")

                        for col_idx, _ in enumerate(data_df.columns):
                            if col_idx == origem_idx or col_idx == destino_idx:
                                continue

                            vehicle = vehicles.iloc[col_idx]
                            viagem = headers.iloc[col_idx]
                            tarifa = data_row.iloc[col_idx]

                            if pd.notna(tarifa) and str(tarifa).strip() != "" and pd.notna(vehicle) and pd.notna(viagem):
                                processed_rows.append({
                                    'Nomeacao': 'N/A',
                                    'Fornecedor': 'N/A',
                                    'Origem': origem,
                                    'LocalColeta': 'N/A',
                                    'Destino': destino,
                                    'DistanciaMin': distancia_min,
                                    'DistanciaMax': distancia_max,
                                    'Transportadora': self._parse_transporter_name(file_name),
                                    'Veiculo': str(vehicle),
                                    'Viagem': str(viagem),
                                    'Tarifa': tarifa,
                                    'Chave': str(origem) + ' & ' + str(destino)
                                })

                    if processed_rows:
                        all_melted_dfs.append(pd.DataFrame(processed_rows))

                except Exception as e:
                    print(f"Error processing file {file_path} for 'FAIXA': {e}")
                    continue

        # !------------------------------FLUXO SPOTS -----------------------------------------!
        elif 'SPOTS' in fluxo_name:
            for file_name in os.listdir(fluxo_path):
                if not file_name.lower().endswith(('.xlsx', '.xls')) or file_name.startswith('~$'):
                    continue
                file_path = os.path.join(fluxo_path, file_name)
                try:
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    sheet = wb.active
                    motorista_cols = {}
                    last_vehicle = None
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_val = sheet.cell(row=1, column=col_idx).value
                        if cell_val and str(cell_val).strip():
                            vehicle_name = str(cell_val).strip()
                            if vehicle_name == '0.75':
                                last_vehicle = '3/4'
                            else:
                                last_vehicle = vehicle_name
                        
                        if last_vehicle:
                            motorista_val_raw = sheet.cell(row=3, column=col_idx).value
                            if motorista_val_raw is not None:
                                try:
                                    motorista_clean = int(float(str(motorista_val_raw).strip()))
                                    motorista_cols[col_idx] = (last_vehicle, motorista_clean)
                                except (ValueError, TypeError):
                                    continue

                    header_map = {}
                    data_start_row, header_found_row = 1, -1
                    for r in range(1, min(10, sheet.max_row + 1)):
                        if header_found_row != -1 and r > header_found_row: break
                        for c in range(1, min(20, sheet.max_column + 1)):
                            cell_val = str(sheet.cell(row=r, column=c).value or '').strip().lower()
                            if 'origem' in cell_val: header_map['Origem'] = c
                            elif 'destino' in cell_val: header_map['Destino'] = c
                        if 'Origem' in header_map or 'Destino' in header_map:
                            header_found_row, data_start_row = r, r + 1
                    
                    if 'Origem' not in header_map or 'Destino' not in header_map: continue
                    
                    processed_rows = []
                    for row_idx in range(data_start_row, sheet.max_row + 1):
                        origem = str(sheet.cell(row=row_idx, column=header_map['Origem']).value or '').strip()
                        destino_raw = str(sheet.cell(row=row_idx, column=header_map['Destino']).value or '').strip()
                        if not origem or not destino_raw: continue

                        # --- CORRECTED PARSING LOGIC ---
                        distancia_min, distancia_max = None, None
                        clean_destino = destino_raw

                        # Scenario 1: "PE 01 KM - 10 Km" or "MG 11-20"
                        match = re.search(r'^(.*?)\s*(\d+)\s*(?:km)?\s*-\s*(\d+)', destino_raw, re.IGNORECASE)
                        if match:
                            clean_destino = match.group(1).strip()
                            distancia_min, distancia_max = int(match.group(2)), int(match.group(3))
                        else:
                            # Scenario 2: "De 21 km a 30 km" -> Destino becomes the same as Origem
                            match = re.search(r'de\s*(\d+)\s*(?:a|-|até)\s*(\d+)', destino_raw, re.IGNORECASE)
                            if match:
                                distancia_min, distancia_max = int(match.group(1)), int(match.group(2))
                                clean_destino = origem
                            else:
                                # Scenario 3: "BA acima 40 km"
                                match = re.search(r'^(.*?)\s*acima (?:de)?\s*(\d+)', destino_raw, re.IGNORECASE)
                                if match:
                                    clean_destino = match.group(1).strip()
                                    distancia_min, distancia_max = int(match.group(2)), float(200)
                                else:
                                    # Scenario 4: "SE até 40 km"
                                    match = re.search(r'^(.*?)\s*até\s*(\d+)', destino_raw, re.IGNORECASE)
                                    if match:
                                        clean_destino = match.group(1).strip()
                                        distancia_min, distancia_max = 0, int(match.group(2))
                        
                        # BUGGY LINE REMOVED: distancia_min, distancia_max = 0,0

                        # --- FINAL CLEANUP & DEFAULTING ---
                        # If no distance range was found after all checks, then we clean and default.
                        if distancia_min is None:
                            # Your logic to handle "MG 21" -> "MG"
                            clean_destino = clean_destino.split(' ')[0].strip()
                            
                            # Set the distances to a default of 0
                            distancia_min = 0
                            distancia_max = 0

                        # Fallback: If parsing results in an empty destination, use Origem.
                        if not clean_destino:
                            clean_destino = origem
                        
                        clean_destino  = clean_destino.split(' ')[0].strip()
                        clean_destino  = clean_destino.split('(')[0].strip()

                        for col_idx, (vehicle, motorista) in motorista_cols.items():
                            tarifa = sheet.cell(row=row_idx, column=col_idx).value
                            if tarifa is not None and str(tarifa).strip() != "":
                                try:
                                    processed_rows.append({
                                        'Transportadora': self._parse_transporter_name(file_name),
                                        'Veiculo': vehicle, 'Motorista': motorista,
                                        'Origem': origem,
                                        'Destino': clean_destino,
                                        'DistanciaMin': distancia_min,
                                        'DistanciaMax': distancia_max,
                                        'Distancia': float(motorista),
                                        'Tarifa': float(tarifa),
                                        'Nomeacao': 'N/A', 'Fornecedor': 'N/A', 'LocalColeta': 'N/A', 'Viagem': 'N/A',
                                        'Chave': f"{origem} & {clean_destino}"
                                    })
                                except (ValueError, TypeError): continue
                    if processed_rows: all_melted_dfs.append(pd.DataFrame(processed_rows))
                except Exception as e:
                    print(f"Error processing file {file_path} for 'SPOTS': {e}")
        else:

            # --- ORIGINAL Logic for '01', '03', etc. ---
            geoship_df = None
            try:
                parent_folder = os.path.dirname(self.base_folder)

                # Look for Geoship table in that parent folder
                geoship_filename = next(
                    (f for f in os.listdir(parent_folder)
                    if 'geoshiptable' in f.lower() and f.endswith(('.xlsx', '.xls'))),
                    None
                )

                if geoship_filename:
                    geoship_full_path = os.path.join(parent_folder, geoship_filename)
                    geoship_df = pd.read_excel(geoship_full_path, engine='openpyxl')
                    geoship_df = geoship_df.rename(columns={
                        'Fornecedor': 'Fornecedor_geoship',
                        'Km Total': 'Distancia_geoship',
                        'Destino Materiais': 'Destino_geoship'
                    })
                    print(f"✅ Successfully loaded data source: '{geoship_filename}'.")
                else:
                    print("⚠️ Warning: 'GeoshipTable' file not found.")

            except FileNotFoundError:
                print(f" Error: Directory not found at '{parent_folder}'.")
            except Exception as e:
                print(f" Error loading GeoshipTable: {e}.")


            # --- ORIGINAL Logic for '01', '03', etc. ---
            for file_name in os.listdir(fluxo_path):
                if not file_name.lower().endswith(('.xlsx', '.xls')) or file_name.startswith('~'):
                    continue
                file_path = os.path.join(fluxo_path, file_name)
                try:
                    header_df = pd.read_excel(file_path, header=None, nrows=2, engine='openpyxl')
                    header_df.iloc[0] = header_df.iloc[0].ffill()
                    new_columns = []
                    for i in range(len(header_df.columns)):
                        top_header = str(header_df.iloc[0, i]).upper().strip()
                        bottom_header = str(header_df.iloc[1, i]).upper().strip()
                        if 'UNNAMED' in top_header or top_header == 'NAN':
                            new_columns.append(bottom_header.lower())
                        elif 'UNNAMED' in bottom_header or bottom_header == 'NAN':
                            new_columns.append(top_header)
                        else:
                            new_columns.append(f"{top_header}_{bottom_header}")

                    df = pd.read_excel(file_path, header=None, skiprows=2, engine='openpyxl')
                    min_cols = min(len(df.columns), len(new_columns))
                    df = df.iloc[:, :min_cols]
                    df.columns = new_columns[:min_cols]

                    df.columns = [col.strip() for col in df.columns]

                    tipo_fluxo_col = next((col for col in df.columns if 'tipo de fluxo' in col.lower()), None)

                    fornecedor_col_name = next((c for c in df.columns if 'fornecedor' in c.lower() and 'codigo' not in c.lower()), None)

                    id_cols_map = {
                        'Nomeacao': next((c for c in df.columns if ('nomeação' in c.lower()) or ('nomeacao' in c.lower())), 'Nomeacao'),
                        'Origem': next((c for c in df.columns if ('cidade de coleta' in c.lower()) or ('cidade_coleta' in c.lower())), 'Origem'),
                        'LocalColeta': next((c for c in df.columns if ('local de coleta' in c.lower()) or ('local_coleta' in c.lower())), 'LocalColeta'),
                        'Destino': next((c for c in df.columns if 'destino materiais' in c.lower()), 'Destino'),
                        'Distancia': next((c for c in df.columns if 'distância' in c.lower()), 'Distancia'),
                    }
                    if fornecedor_col_name:
                        id_cols_map['Fornecedor'] = fornecedor_col_name

                    df.rename(columns={v: k for k, v in id_cols_map.items() if v in df.columns}, inplace=True)

                    if 'Fornecedor' not in df.columns:
                        df['Fornecedor'] = 'N/A'

                    id_vars = list(id_cols_map.keys())
                    if tipo_fluxo_col:
                        id_vars.append(tipo_fluxo_col)

                    df['Transportadora'] = self._parse_transporter_name(file_name)
                    id_vars.append('Transportadora')

                    value_vars = [col for col in df.columns if '_' in col and col not in id_vars]

                    melted_df = df.melt(
                        id_vars=[v for v in id_vars if v in df.columns],
                        value_vars=value_vars,
                        var_name='Veiculo_Viagem',
                        value_name='Tarifa'
                    )

                    melted_df[['Veiculo', 'Viagem']] = melted_df['Veiculo_Viagem'].str.split('_', expand=True, n=1)
                    melted_df.drop('Veiculo_Viagem', axis=1, inplace=True)
                    melted_df['Chave'] = melted_df['Origem'].astype(str) + ' & ' + melted_df['Destino'].astype(str)

                    # Replace Geoship Rows
                    # Replace Geoship Rows
                    if tipo_fluxo_col and tipo_fluxo_col in melted_df.columns and geoship_df is not None:
                        is_geoship = melted_df[tipo_fluxo_col].astype(str).str.lower().str.contains('geoship', na=False)
                        geoship_matches = melted_df[is_geoship]

                        non_geoship = melted_df[~is_geoship]
                        updated_rows = []

                        # For each row where Tipo de Fluxo contains Geoship
                        for _, row in geoship_matches.iterrows():
                            tipo_fluxo_value = str(row[tipo_fluxo_col]).strip()
                            # Search Geoship table for rows where Tipo de Fluxo matches (or contains) the tipo_fluxo_value
                            # Here we assume geoship_df has a column like 'Tipo de Fluxo' or similar to match
                            # Adjust the column name below as needed, for example 'Tipo de Fluxo' or 'Tipo_de_Fluxo' or something else
                            # We'll assume it's 'Tipo de Fluxo' exactly for now
                            geoship_key_col = next((col for col in geoship_df.columns if 'tipo' in col.lower() and 'fluxo' in col.lower()), None)
                            if geoship_key_col is None:
                                # fallback: just match on some column named 'Geoship' if exists
                                geoship_key_col = next((col for col in geoship_df.columns if 'geoship' in col.lower()), None)
                            
                            if geoship_key_col:
                                # Find rows in geoship_df that contain the tipo_fluxo_value (case insensitive)
                                matched_geo_rows = geoship_df[
                                            geoship_df[geoship_key_col].astype(str).str.lower() == tipo_fluxo_value.lower()
                                        ]

                            else:
                                matched_geo_rows = pd.DataFrame()  # no matches possible

                            if matched_geo_rows.empty:
                                # No match found, keep original row as is
                                updated_rows.append(row)
                            else:
                                # For each matching geoship row, create a new combined row
                                for _, geo_row in matched_geo_rows.iterrows():
                                    new_row = row.copy()
                                    # Update columns with geoship data, rename accordingly if needed
                                    # Match columns you renamed earlier: Fornecedor_geoship, Distancia_geoship, Destino_geoship, plus others if relevant
                                    # You can add any other columns you want to override from geo_row here
                                    new_row['Fornecedor'] = geo_row.get('Fornecedor_geoship', new_row.get('Fornecedor', 'N/A'))
                                    new_row['Distancia'] = geo_row.get('Distancia_geoship', new_row.get('Distancia', None))
                                    new_row['Origem'] = geo_row.get('CNPJ Origem', new_row.get('Origem', None))
                                    new_row['Destino'] = geo_row.get('Destino_geoship', new_row.get('Destino', None))
                                    # You can add other columns from geo_row if needed

                                    updated_rows.append(new_row)

                        # Combine all non-geoship rows and expanded geoship rows
                        melted_df = pd.concat([non_geoship, pd.DataFrame(updated_rows)], ignore_index=True).drop(columns=[tipo_fluxo_col], errors='ignore')



                    all_melted_dfs.append(melted_df)

                except Exception as e:
                    print(f"Error processing file {file_path}: {e}")
                    continue


        
        if not all_melted_dfs:
            return {'success': False, 'message': 'Nenhum arquivo de tarifa válido foi encontrado.'}

        master_df = pd.concat(all_melted_dfs, ignore_index=True).dropna(subset=['Tarifa'])
        if 'Veiculo' in master_df.columns:
            master_df['Veiculo'] = master_df['Veiculo'].apply(normalize_vehicle_name)
        # --- NEW & IMPROVED DATA CLEANING ---
        str_cols = ['Nomeacao', 'Fornecedor', 'Origem', 'Local Coleta', 'Destino', 'Veiculo', 'Viagem']
        for col in str_cols:
            if col in master_df.columns:
                master_df[col] = master_df[col].astype(str).str.strip()

        if 'Viagem' in master_df.columns:
            master_df['Viagem'] = master_df['Viagem'].str.upper()

        special_cols = ['Nomeacao', 'Fornecedor']
        for col in special_cols:
             if col in master_df.columns:
                master_df[col] = master_df[col].str.upper().str.replace(' ', '').fillna('N/A')

        if 'Nomeacao' in master_df.columns:
            master_df['Nomeacao'] = master_df['Nomeacao'].replace('PRINCIPALCARRETA', 'PRINCIPAL')
        # --- END OF CLEANING ---

        for col in ['Tarifa', 'Distancia']:
            if col in master_df.columns:
                master_df[col] = pd.to_numeric(master_df[col], errors='coerce')

        filters = {
            'Nomeacao': sorted(master_df['Nomeacao'].dropna().unique().tolist()),
            'Fornecedores': sorted(master_df['Fornecedor'].dropna().unique().tolist()),
            'Origem': sorted(master_df['Origem'].dropna().unique().tolist()),
            'LocaisColeta': sorted(master_df['LocalColeta'].dropna().unique().tolist()),
            'Destino': sorted(master_df['Destino'].dropna().unique().tolist()),
            'Veiculos': sorted(master_df['Veiculo'].dropna().unique().tolist())
        }

        # **[FIX]** More robustly create the Motoristas filter list
        if 'Motorista' in master_df.columns:
            try:
                # Ensure values are clean integers before creating the unique list
                unique_motoristas = master_df['Motorista'].dropna().astype(int).unique()
                filters['Motoristas'] = sorted(unique_motoristas.tolist())
            except (ValueError, TypeError):
                # If conversion fails for any reason, return an empty list
                filters['Motoristas'] = []
        
        response = {'success': True, 'filters': filters}
        self.fluxo_data[fluxo_name] = {'df': master_df, 'filters_response': response}
        
        return response

   
    # (Inside your Api class)
    def calculate_tariffs(self, params):
        fluxo_name = params['fluxo']
        if not fluxo_name or fluxo_name not in self.fluxo_data: return []
        df = self.fluxo_data[fluxo_name]['df'].copy()
        
        # --- Apply Filters ---
        if params.get('fornecedor'): df = df[df['Fornecedor'] == params['fornecedor']]
        if params.get('nomeacao'): df = df[df['Nomeacao'] == params['nomeacao']]
        if params.get('origem'): df = df[df['Origem'] == params['origem']]
        if params.get('local_coleta') and 'LocalColeta' in df.columns: df = df[df['LocalColeta'] == params['local_coleta']]
        if params.get('destino'): df = df[df['Destino'] == params['destino']]
        if params.get('veiculo'): df = df[df['Veiculo'] == params['veiculo']]
        
        if 'SPOTS' not in fluxo_name.upper():
            if params.get('calc_type'): 
                df = df[df['Viagem'] == params['calc_type']]

        if params.get('motorista') and 'Motorista' in df.columns:
            try: df = df[df['Motorista'] == int(params['motorista'])]
            except (ValueError, TypeError): pass

        if 'Tarifa' in df.columns:
            df = df[pd.to_numeric(df['Tarifa'], errors='coerce').fillna(0) > 0]

        if df.empty: return []

        # --- Perform Calculations ---
        try:
            new_distance = float(params.get('km_value'))
            if new_distance <= 0: new_distance = None
        except (ValueError, TypeError): new_distance = None

        if new_distance:
            df['Tarifa Real'] = pd.NA
            
            is_range_based_fluxo = ('FAIXA' in fluxo_name.upper() or 'MILK RUN' in fluxo_name.upper() or 'SPOTS' in fluxo_name.upper()) and 'DistanciaMin' in df.columns

            if is_range_based_fluxo:
                range_mask = (df['DistanciaMin'] <= new_distance) & (df['DistanciaMax'] >= new_distance) & (df['DistanciaMax'] > 0)
                
                if range_mask.any():
                    if 'MILK RUN' in fluxo_name.upper():
                        df.loc[range_mask, 'Tarifa Real'] = new_distance * df.loc[range_mask, 'Tarifa']
                    elif 'SPOTS' in fluxo_name.upper():
                        df.loc[range_mask, 'Tarifa Real'] = (new_distance * df.loc[range_mask, 'Tarifa']) / df.loc[range_mask, 'Distancia']
                        df.loc[range_mask, 'Distancia'] = new_distance
                    else: # FAIXA
                        df.loc[range_mask, 'Tarifa Real'] = df.loc[range_mask, 'Tarifa']
                    
                    df = df[range_mask].copy()
                else:
                    if 'Distancia' in df.columns:
                        fallback_mask = (df['DistanciaMin'] == 0) & (df['DistanciaMax'] == 0) & (df['Distancia'] > 0)
                        df.loc[fallback_mask, 'Tarifa Real'] = (new_distance * df.loc[fallback_mask, 'Tarifa']) / df.loc[fallback_mask, 'Distancia']
                        df.loc[fallback_mask, 'Distancia'] = new_distance
                        df = df[fallback_mask].copy()
                    else:
                        df = df.iloc[0:0]

            elif 'Distancia' in df.columns:
                # --- START OF FIX ---
                # Before calculating, find the single best (cheapest) tariff for each unique route to avoid duplicates.
                # A unique route is defined by Origin, Destination, Vehicle, and Transporter.
                key_cols = ['Origem', 'Destino', 'Veiculo', 'Transportadora']

                # Ensure all key columns exist before de-duplicating
                if all(col in df.columns for col in key_cols):
                    df = df.sort_values('Tarifa', ascending=True).drop_duplicates(subset=key_cols, keep='first')
                # --- END OF FIX ---

                mask = (df['Distancia'].notna()) & (df['Distancia'] > 0)
                df.loc[mask, 'Tarifa Real'] = (new_distance * df.loc[mask, 'Tarifa']) / df.loc[mask, 'Distancia']
                df.loc[mask, 'Distancia'] = new_distance
                
        else:
            if 'SPOTS' in fluxo_name.upper() and 'Distancia' in df.columns:
                df['Distancia'] = pd.NA

        if 'Tarifa Real' in df.columns:
            df = df[df['Tarifa Real'] > 0]

        # --- Final Processing and Return ---
        sort_col = 'Tarifa Real' if 'Tarifa Real' in df.columns and df['Tarifa Real'].notna().any() else 'Tarifa'
        df_sorted = df.sort_values(by=sort_col, ascending=True)

        base_display_cols = ['Origem','Destino','Veiculo','Motorista','Transportadora','DistanciaMin','DistanciaMax','Distancia','Tarifa']
        final_display_cols = base_display_cols.copy()
        if new_distance:
            final_display_cols.append('Tarifa Real')

        existing_cols = [col for col in final_display_cols if col in df_sorted.columns]
        self.last_results_df = df_sorted[existing_cols].copy()

        self.last_results_df.rename(columns={'DistanciaMin': 'Dist. Mín','DistanciaMax': 'Dist. Máx','Transportadora': 'Transp.'}, inplace=True)

        if 'Tarifa Real' in self.last_results_df.columns:
            self.last_results_df['Tarifa Real'] = pd.to_numeric(self.last_results_df['Tarifa Real'], errors='coerce').round(2)
        if 'Tarifa' in self.last_results_df.columns:
            self.last_results_df['Tarifa'] = pd.to_numeric(self.last_results_df['Tarifa'], errors='coerce').round(2)
        
        df_for_json = self.last_results_df.replace([np.inf, -np.inf, np.nan], None)
        
        return df_for_json.to_dict('records')


    def export_to_excel(self):
        if self.last_results_df is None or self.last_results_df.empty:
            return {'success': False, 'message': 'Não há dados para exportar.'}
        if not self._window:
            return {'success': False, 'message': 'Erro: Referência da janela não encontrada.'}
        try:
            file_path_tuple = self._window.create_file_dialog(webview.SAVE_DIALOG, directory=os.path.expanduser('~'), save_filename='cotacao_tarifas.xlsx')
            if not file_path_tuple:
                return {'success': False, 'message': 'Exportação cancelada.'}
            user_choice = file_path_tuple[0]
            filename = os.path.basename(user_choice)
            if not filename or ":" in filename:
                downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
                os.makedirs(downloads_path, exist_ok=True)
                final_path = os.path.join(downloads_path, 'cotacao_tarifas.xlsx')
            else:
                final_path = user_choice
                if not final_path.lower().endswith('.xlsx'): final_path += '.xlsx'
            self.last_results_df.to_excel(final_path, index=False, engine='openpyxl')
            return {'success': True, 'message': f'Sucesso! Arquivo salvo em: {final_path}'}
        except Exception as e:
            return {'success': False, 'message': f'Erro ao exportar: {e}'}
if __name__ == '__main__':
    api = Api()

    window = webview.create_window(
        'Calculador de Tarifas Unificado',
        resource_path('gui.html'),  # CORRECT: Pass the file path directly
        js_api=api,
        width=1400,
        height=900,
        resizable=True
    )
    
    api._window = window

    # Add debug=True here to help find any other issues
    webview.start(debug=False)