import webview
import json
import os
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import openpyxl

# --- Front-End User Interface (gui.html) ---
HTML_CONTENT = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Calculador de Tarifas Unificado</title>
    <script src="https://cdn.tailwindcss.com"></script>
    
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/choices.js/public/assets/styles/choices.min.css"/>
    <script src="https://cdn.jsdelivr.net/npm/choices.js/public/assets/scripts/choices.min.js"></script>
    
    <style>
        html, body { height: 100%; overflow: hidden; }
        body { display: flex; flex-direction: column; }
        @keyframes fadeIn { from { opacity: 0; transform: translateY(-10px); } to { opacity: 1; transform: translateY(0); } }
        .fade-in { animation: fadeIn 0.5s ease-out forwards; }
        
        ::-webkit-scrollbar { width: 8px; height: 8px;}
        ::-webkit-scrollbar-track { background: #f1f1f1; border-radius: 10px;}
        ::-webkit-scrollbar-thumb { background: #888; border-radius: 10px;}
        ::-webkit-scrollbar-thumb:hover { background: #555; }

        #notification { transition: opacity 0.5s, transform 0.5s; transform: translateY(20px); max-width: 90%; }
        
        #results-container-wrapper { overflow: auto; }
        #results-table th { position: sticky; top: 0; background-color: #e5e7eb; z-index: 1; }
        #results-table td, #results-table th {
            user-select: text; /* or user-select: all; */
            cursor: text;
        }
        
        .control-grid {
            display: grid;
            grid-template-columns: 1fr;
            gap: 1rem;
        }
        select, input[type=number] {
            border-radius: 0.5rem;
        }

        /* Custom styles for Choices.js to match the UI */
        .choices { width: 100%; margin-bottom: 0; }
        .choices__inner {
            background-color: #fff;
            border-radius: 0.5rem;
            border: 1px solid rgb(209 213 219);
            padding: 0.45rem 0.75rem;
            font-size: 0.9rem;
            min-height: auto;
        }
        .is-open .choices__inner { border-radius: 0.5rem 0.5rem 0 0; }
        .choices__list--dropdown { border-radius: 0 0 0.5rem 0.5rem; }
        .choices[data-type*="select-one"]::after { right: 12.5px; }
        .choices.is-disabled .choices__inner, .choices.is-disabled .choices__input {
            background-color: #f3f4f6;
            cursor: not-allowed;
            -webkit-user-select: none;
            user-select: none;
        }

        option{
        font-size: 0.9rem;
        }
    </style>
</head>
<body class="bg-slate-100 font-sans">
    <div class="flex flex-col h-screen bg-white">
        <header class="sticky top-0 bg-indigo-900 shadow-md z-20 p-3">
            <div class="w-full px-6 flex justify-between items-center">
                <div class="flex items-center gap-4">
                    <button id="menu-toggle-btn" onclick="toggleSidebar()" class="text-white p-2 rounded-md hover:bg-indigo-700">
                        <svg xmlns="http://www.w3.org/2000/svg" class="h-6 w-6" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                            <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 6h16M4 12h16m-7 6h7" />
                        </svg>
                    </button>
                    <div>
                        <h1 class="text-xl font-bold text-white">Calculadora de Tarifas Por Fluxo</h1>
                        <p id="folder-path" class="text-xs text-indigo-300">Nenhuma pasta principal selecionada.</p>
                    </div>
                </div>
                <button onclick="selectBaseFolder()" class="bg-indigo-600 text-white font-bold py-2 px-4 rounded-lg shadow-md hover:bg-indigo-700 transition flex-shrink-0">
                    Selecionar Pasta Principal
                </button>
            </div>
        </header>

        <main class="flex-grow overflow-auto p-6 w-full">
            <div class="flex flex-row gap-6 h-full">
                <div id="sidebar" class="lg:w-70 flex-shrink-0 bg-slate-50 p-3 rounded-lg shadow-sm flex flex-col gap-4 transition-all duration-300">
                    <div class="control-grid">
                        <div>
                            <label for="fluxo-select" class="block text-sm font-medium text-slate-700 mb-2">Escolher Fluxo</label>
                            <select id="fluxo-select" onchange="onFluxoSelected()" class="w-full p-1 bg-white border font-sm " disabled>
                                <option>Selecione uma pasta</option>
                            </select>
                        </div>

                        
                        <div>
                            <label for="nomeacao-select" class="block text-sm font-medium text-slate-700 mb-2">Nomeação</label>
                            <select id="nomeacao-select" class="w-full p-1 bg-white border font-sm text-sm" disabled>
                                <option>Selecione um fluxo</option>
                            </select>
                        </div>
                        <div>
                            <label for="fornecedor-select" class="block text-sm font-medium text-slate-700 mb-2">Fornecedor</label>
                            <select id="fornecedor-select" class="w-full p-1 bg-white border" disabled>
                                <option>Selecione um fluxo</option>
                            </select>
                        </div>
                        <div>
                            <label for="origem-select" class="block text-sm font-medium text-slate-700 mb-2">Cidade de Coleta</label>
                            <select id="origem-select" class="w-full p-1 bg-white border" disabled>
                                <option>Selecione um fluxo</option>
                            </select>
                        </div>
                        <div>
                            <label for="local-coleta-select" class="block text-sm font-medium text-slate-700 mb-2">Local de Coleta</label>
                            <select id="local-coleta-select" class="w-full p-1 bg-white border" disabled>
                                <option>Selecione um fluxo</option>
                            </select>
                        </div>
                        <div>
                            <label for="destino-select" class="block text-sm font-medium text-slate-700 mb-2">Destino Materiais</label>
                            <select id="destino-select" class="w-full p-1 bg-white border" disabled>
                                <option>Selecione um fluxo</option>
                            </select>
                        </div>
                        <div>
                            <label for="vehicle-select" class="block text-sm font-medium text-slate-700 mb-2">Tipo de Veículo</label>
                            <select id="vehicle-select" class="w-full p-1 bg-white border" disabled>
                                <option>Selecione um fluxo</option>
                            </select>
                        </div>

                        <div id="motorista-filter-div" class="hidden">
                            <label for="motorista-select" class="block text-sm font-medium text-slate-700 mb-2">Motorista(s)</label>
                            <select id="motorista-select" class="w-full p-1 bg-white border" disabled>
                                <option>Selecione as opções</option>
                            </select>
                        </div>

                        <div>
                            <label class="block text-sm font-medium text-slate-700 mb-2">Viagem</label>
                            <div class="flex items-center justify-between bg-white border rounded-lg p-1">
                                <button id="calc-type-OW" onclick="setCalcType('OW')" class="w-1/2 py-2 text-sm font-semibold bg-indigo-600 text-white rounded-md">One Way (OW)</button>
                                <button id="calc-type-RT" onclick="setCalcType('RT')" class="w-1/2 py-2 text-sm font-semibold text-slate-600 rounded-md">Round Trip (RT)</button>
                            </div>
                        </div>
                        <div>
                            <label for="km-value" class="block text-sm font-medium text-slate-700 mb-2">Nova Distância (KM) (Opcional)</label>
                            <input type="number" id="km-value" placeholder="Para 'Regra de Três'" class="w-full p-1 bg-white border">
                        </div>
                    </div>
                    <div class="mt-auto pt-2">
                        <button onclick="performCalculation()" class="w-full bg-blue-600 text-white font-bold p-3 rounded-lg shadow-md hover:bg-blue-700">
                            Calcular Tarifas
                        </button>
                    </div>
                </div>

                <div class="flex-grow flex flex-col">
                    <div id="calc-placeholder" class="flex-grow h-full flex flex-col items-center justify-center text-center py-5 px-2 border-2 border-dashed border-slate-300 rounded-lg">
                        <svg xmlns="http://www.w3.org/2000/svg" class="w-24 h-24 text-slate-300 mb-4" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1" stroke-linecap="round" stroke-linejoin="round">
                           <path d="M14 17.5V19a2 2 0 0 1-2 2H4a2 2 0 0 1-2-2V7a2 2 0 0 1 2-2h12v2.5"/>
                           <path d="M12 17.5h8.5l3-4.5-3-4.5H12Z"/><circle cx="6.5" cy="17.5" r="2.5"/><circle cx="18.5" cy="17.5" r="2.5"/>
                        </svg>
                        <p class="text-slate-500">Os resultados da cotação aparecerão aqui.</p>
                    </div>
                    <div id="calc-loading" class="flex-grow h-full hidden items-center justify-center text-center py-100"><p class="text-slate-600">Calculando...</p></div>
                    
                    <div id="results-table" class="hidden h-full flex flex-col">
                         <div class="flex justify-between items-center mb-1 flex-shrink-0">
                             <h2 class="text-xl font-bold text-slate-700">Resultados da Tarifas </h2>
                             <button id="export-btn" onclick="exportResults()" class="bg-green-600 text-white font-bold py-2 px-4 rounded-lg shadow-md hover:bg-green-700">
                                 Exportar para Excel
                             </button>
                         </div>
                         <div id="results-container-wrapper" class="border rounded-lg flex-grow">
                             <table class="min-w-full bg-white"><thead class="bg-slate-200" id="results-head"></thead><tbody id="results-body" class="divide-y divide-slate-200"></tbody></table>
                         </div>
                    </div>
                </div>
            </div>
        </main>
        
        <footer class="bg-white border-t p-2 z-10"><div class="w-full px-6 flex justify-between items-center"><p class="text-xs text-slate-400">Desenvolvido por Vincent Pernah - Analista de Projetos</p><div class="flex items-center gap-4"><span class="font-bold text-sm text-slate-700">Stellantis</span><span class="font-bold text-sm" style="color: #D40511;">DHL</span></div></div></footer>
    </div>

    <div id="notification" class="fixed bottom-12 right-5 text-white py-2 px-4 rounded-lg shadow-lg opacity-0 pointer-events-none z-30"></div>

    
# !------------------------------ Java Script Contions and Code -----------------------------------------!
    <script>
        let currentCalcType = 'OW';
        let choiceInstances = {};

        function toggleSidebar() {
            const sidebar = document.getElementById('sidebar');
            sidebar.classList.toggle('hidden');
        }

        function destroyAllChoiceInstances() {
            for (const id in choiceInstances) {
                if (choiceInstances[id]) {
                    choiceInstances[id].destroy();
                    delete choiceInstances[id];
                }
            }
        }

        function initializeSearchableDropdown(elementId) {
            const element = document.getElementById(elementId);
            if (element) {
                const choices = new Choices(element, {
                    searchEnabled: true,
                    itemSelectText: '✔️',
                    noResultsText: 'Nenhum resultado',
                    noChoicesText: 'Sem opções'
                });
                choiceInstances[elementId] = choices;
            }
        }

        function setCalcType(type) {
            currentCalcType = type;
            const btnOW = document.getElementById('calc-type-OW');
            const btnRT = document.getElementById('calc-type-RT');
            
            if (type === 'OW') {
                btnOW.classList.add('bg-indigo-600', 'text-white');
                btnOW.classList.remove('text-slate-600');
                btnRT.classList.remove('bg-indigo-600', 'text-white');
                btnRT.classList.add('text-slate-600');
            } else { // RT
                btnRT.classList.add('bg-indigo-600', 'text-white');
                btnRT.classList.remove('text-slate-600');
                btnOW.classList.remove('bg-indigo-600', 'text-white');
                btnOW.classList.add('text-slate-600');
            }
        }

        async function selectBaseFolder() {
            try {
                const response = await window.pywebview.api.select_folder();
                if (response.success) {
                    document.getElementById('folder-path').textContent = `Pasta: ${response.path}`;
                    showNotification('Pasta principal selecionada!', true);
                    populateDropdown('fluxo-select', response.fluxos, 'Selecione um fluxo');
                    document.getElementById('fluxo-select').disabled = false;
                    resetAllFilters();
                } else {
                    showNotification(response.message, false);
                }
            } catch (e) {
                showNotification('Ocorreu um erro ao selecionar a pasta.', false);
            }
        }

        async function onFluxoSelected() {
            const fluxo = document.getElementById('fluxo-select').value;
            const motoristaFilterDiv = document.getElementById('motorista-filter-div');

            // Show/hide motorista filter based on fluxo
            if (fluxo && fluxo.toUpperCase().includes('SPOTS')) {
                motoristaFilterDiv.classList.remove('hidden');
            } else {
                motoristaFilterDiv.classList.add('hidden');
            }

            if (!fluxo) {
                resetAllFilters();
                return;
            }
            
            showLoadingOnFilters();

            try {
                const response = await window.pywebview.api.get_filters_for_fluxo(fluxo);
                if (response.success) {
                    populateDropdown('nomeacao-select', response.filters.Nomeacao, 'Todas');
                    populateDropdown('fornecedor-select', response.filters.Fornecedores, 'Todos');
                    populateDropdown('origem-select', response.filters.Origem, 'Todas');
                    populateDropdown('local-coleta-select', response.filters.LocaisColeta, 'Todos');
                    populateDropdown('destino-select', response.filters.Destino, 'Todos');
                    populateDropdown('vehicle-select', response.filters.Veiculos, 'Todos');
                    
                    // Populate motorista dropdown if it's a SPOTS flow
                    if (fluxo.toUpperCase().includes('SPOTS')) {
                        populateDropdown('motorista-select', response.filters.Motoristas, 'Todos');
                        initializeSearchableDropdown('motorista-select');
                    }

                    initializeSearchableDropdown('fornecedor-select');
                    initializeSearchableDropdown('origem-select');
                    initializeSearchableDropdown('local-coleta-select');
                    initializeSearchableDropdown('destino-select');
                    initializeSearchableDropdown('vehicle-select');

                    enableAllFilters();
                } else {
                    showNotification(response.message, false);
                    resetAllFilters();
                }
            } catch (e) {
                showNotification('Erro ao carregar filtros para este fluxo: ' + e, false);
                resetAllFilters();
            }
        }

        async function performCalculation() {
            const params = {
                fluxo: document.getElementById('fluxo-select').value,
                nomeacao: document.getElementById('nomeacao-select').value,
                fornecedor: document.getElementById('fornecedor-select').value,
                origem: document.getElementById('origem-select').value,
                local_coleta: document.getElementById('local-coleta-select').value,
                destino: document.getElementById('destino-select').value,
                veiculo: document.getElementById('vehicle-select').value,
                motorista: document.getElementById('motorista-select').value,
                calc_type: currentCalcType,
                km_value: document.getElementById('km-value').value
            };

            if (!params.fluxo) {
                showNotification('Por favor, selecione um fluxo.', false);
                return;
            }

            document.getElementById('calc-loading').style.display = 'flex';
            document.getElementById('calc-placeholder').style.display = 'none';
            document.getElementById('results-table').classList.add('hidden');
            
            try {
                const results = await window.pywebview.api.calculate_tariffs(params);
                displayCalcResults(results);
            } catch (e) {
                showNotification("Ocorreu um erro ao calcular.", false);
                console.error(e);
            } finally {
                document.getElementById('calc-loading').style.display = 'none';
            }
        }

        function displayCalcResults(results) {
            const head = document.getElementById('results-head');
            const body = document.getElementById('results-body');
            const resultsTable = document.getElementById('results-table');
            const placeholder = document.getElementById('calc-placeholder');
            
            head.innerHTML = '';
            body.innerHTML = '';

            if (!results || results.length === 0) {
                placeholder.style.display = 'flex';
                placeholder.innerHTML = `<p class="text-orange-600">Nenhum resultado encontrado para os critérios selecionados.</p>`;
                resultsTable.classList.add('hidden');
                return;
            }

            const headers = Object.keys(results[0]);
            const headerRow = document.createElement('tr');
            headers.forEach(h => {
                const th = document.createElement('th');
                th.className = 'text-left text-sm font-bold py-3 px-4 whitespace-nowrap';
                th.textContent = h;
                headerRow.appendChild(th);
            });
            head.appendChild(headerRow);

            results.forEach(result => {
                const row = document.createElement('tr');
                row.className = 'hover:bg-slate-50 fade-in';
                headers.forEach(header => {
                    const td = document.createElement('td');
                    td.className = 'py-2 px-4 text-sm';
                    
                    if (header === 'Tarifa' && typeof result[header] === 'number') {
                        td.textContent = result[header].toLocaleString('pt-BR', { style: 'currency', currency: 'BRL' });
                    } else {
                        td.textContent = result[header];
                    }
                    row.appendChild(td);
                });
                body.appendChild(row);
            });
            
            resultsTable.classList.remove('hidden');
            placeholder.style.display = 'none';
        }

        async function exportResults() {
            try {
                showNotification('Abrindo diálogo para salvar...', true);
                const response = await window.pywebview.api.export_to_excel();
                if (response.message !== 'Exportação cancelada.') {
                    showNotification(response.message, response.success);
                }
            } catch (e) {
                showNotification('Falha ao exportar.', false);
            }
        }
        
        function populateDropdown(id, items, defaultText) {
            const select = document.getElementById(id);
            select.innerHTML = `<option value="">${defaultText}</option>`;
            (items || []).forEach(item => {
                const option = document.createElement('option');
                option.value = item;
                option.textContent = item;
                select.appendChild(option);
            });
        }
        
        function resetAllFilters() {
            destroyAllChoiceInstances();
            ['nomeacao-select', 'fornecedor-select', 'origem-select', 'local-coleta-select', 'destino-select', 'vehicle-select', 'motorista-select'].forEach(id => {
                const select = document.getElementById(id);
                if (select) {
                    select.innerHTML = '<option>Selecione um fluxo</option>';
                    select.disabled = true;
                }
            });
            document.getElementById('motorista-filter-div').classList.add('hidden');
        }

        function showLoadingOnFilters() {
            destroyAllChoiceInstances();
             ['nomeacao-select', 'fornecedor-select', 'origem-select', 'local-coleta-select', 'destino-select', 'vehicle-select', 'motorista-select'].forEach(id => {
                const select = document.getElementById(id);
                if (select) {
                    select.innerHTML = '<option>Carregando...</option>';
                    select.disabled = true;
                }
            });
        }
        
        function enableAllFilters() {
             const filterIds = ['nomeacao-select', 'fornecedor-select', 'origem-select', 'local-coleta-select', 'destino-select', 'vehicle-select', 'motorista-select'];
             filterIds.forEach(id => {
                 const instance = choiceInstances[id];
                 if (instance) {
                        instance.enable();
                 } else {
                        const element = document.getElementById(id);
                        if (element) {
                            element.disabled = false;
                        }
                 }
            });
        }

        function showNotification(message, isSuccess) {
            const notification = document.getElementById('notification');
            notification.textContent = message;
            notification.className = `fixed bottom-12 right-5 text-white py-2 px-4 rounded-lg shadow-lg opacity-0 pointer-events-none z-30 ${isSuccess ? 'bg-green-600' : 'bg-red-600'}`;
            notification.classList.remove('opacity-0');
            notification.style.transform = 'translateY(0)';
            setTimeout(() => {
                notification.classList.add('opacity-0');
                notification.style.transform = 'translateY(20px)';
            }, 5000);
        }
    </script>
</body>
</html>
"""


























# !------------------------------ API code and conditions -----------------------------------------!
class Api:
    def __init__(self):
        self.base_folder = None
        self.fluxo_data = {}
        self.last_results_df = None
        self._window = None


    def _parse_transporter_name(self, filename):
        """Extracts a cleaned-up transporter name from the filename."""
        match = re.search(r'_(.*?)\.', filename)
        if match:
            return match.group(1).replace('_', ' ').title()
        return os.path.splitext(filename)[0].replace('_', ' ').title()

    def select_folder(self):
        """Handles the main folder selection using pywebview's native dialog."""
        try:
            if not self._window:
                return {'success': False, 'message': 'Erro: A referência da janela não foi configurada.'}
            
            dialog_result = self._window.create_file_dialog(webview.FOLDER_DIALOG)

            if not dialog_result:
                return {'success': False, 'message': 'Nenhuma pasta foi selecionada.'}
            
            folder_path = dialog_result[0]
            self.base_folder = folder_path
            self.fluxo_data = {}
            
            fluxos = [d for d in os.listdir(self.base_folder) if os.path.isdir(os.path.join(self.base_folder, d))]
            
            if not fluxos:
                return {'success': False, 'message': 'Nenhum fluxo (subpasta) encontrado na pasta selecionada.'}
                
            return {'success': True, 'path': self.base_folder, 'fluxos': sorted(fluxos)}
            
        except Exception as e:
            print(f"PYTHON ERROR in select_folder: {e}") 
            return {'success': False, 'message': f'Ocorreu um erro no Python: {e}'}

    def get_filters_for_fluxo(self, fluxo_name):
        """
        Reads all Excel files in a given fluxo, processes them into a tidy (long) DataFrame,
        and returns the available filters for the UI.
        """
        if fluxo_name in self.fluxo_data:
            return self.fluxo_data[fluxo_name]['filters_response']

        fluxo_path = os.path.join(self.base_folder, fluxo_name)
        all_melted_dfs = []


        if '04. MILK RUN' in fluxo_name:
            for file_name in os.listdir(fluxo_path):
                if not file_name.lower().endswith(('.xlsx', '.xls')) or file_name.startswith('~$'):
                    continue
                
                file_path = os.path.join(fluxo_path, file_name)
                try:
                    full_df = pd.read_excel(file_path, header=None, engine='openpyxl', keep_default_na=False, na_values=[''])

                    header_row_idx = -1
                    faixa_col_idx = -1
                    
                    for r_idx in range(min(10, len(full_df))):
                        row_as_str = full_df.iloc[r_idx].astype(str).str.strip().str.upper()
                        matches = row_as_str[row_as_str == 'FAIXA KM']
                        if not matches.empty:
                            header_row_idx = r_idx
                            faixa_col_idx = matches.index[0]
                            break
                    
                    if header_row_idx == -1:
                        print(f"Skipping file {file_name}: Could not find 'FAIXA KM' header.")
                        continue

                    rt_info_row_idx = header_row_idx - 1
                    data_start_row_idx = header_row_idx + 1

                    if rt_info_row_idx < 0 or data_start_row_idx >= len(full_df):
                        print(f"Skipping file {file_name}: Invalid structure around 'FAIXA KM' header.")
                        continue

                    # FIX 1: Add .ffill() to carry over the trip type to all vehicles
                    round_trip_info = full_df.iloc[rt_info_row_idx].ffill()
                    vehicle_headers = full_df.iloc[header_row_idx]
                    data_df = full_df.iloc[data_start_row_idx:]
                    
                    vehicle_col_indices = [
                        idx for idx, val in enumerate(vehicle_headers)
                        if idx > faixa_col_idx and pd.notna(val) and str(val).strip() != ""
                    ]

                    processed_rows = []
                    for _, data_row in data_df.iterrows():
                        faixa_km_str = str(data_row.iloc[faixa_col_idx]).strip()
                        distancia_min = distancia_max = None

                        match = re.search(r'(\d+)\s*(?:a|-|até)\s*(\d+)', faixa_km_str, re.IGNORECASE)
                        if match:
                            distancia_min, distancia_max = int(match.group(1)), int(match.group(2))
                        else:
                            match = re.search(r'acima de\s*(\d+)', faixa_km_str, re.IGNORECASE)
                            if match:
                                distancia_min, distancia_max = int(match.group(1)), float('inf')
                            else:
                                match = re.search(r'até\s*(\d+)', faixa_km_str, re.IGNORECASE)
                                if match:
                                    distancia_min, distancia_max = 0, int(match.group(1))

                        if distancia_min is None:
                            continue

                        for col_idx in vehicle_col_indices:
                            try:
                                vehicle = vehicle_headers.iloc[col_idx]
                                viagem = round_trip_info.iloc[col_idx]
                                tarifa = data_row.iloc[col_idx]

                                if pd.notna(tarifa) and str(tarifa).strip() not in ("", "nan") and pd.notna(vehicle) and pd.notna(viagem):
                                    viagem_str = str(viagem).upper().strip()
                                    if 'ROUND' in viagem_str:
                                        viagem_code = 'RT'
                                    elif 'ONE WAY' in viagem_str or 'OW' in viagem_str:
                                        viagem_code = 'OW'
                                    else:
                                        viagem_code = viagem_str
                                    
                                    processed_rows.append({
                                        'Nomeacao': 'N/A', 'Fornecedor': 'N/A', 'Origem': 'N/A',
                                        'LocalColeta': 'N/A', 'Destino': 'N/A',
                                        'DistanciaMin': distancia_min,
                                        'DistanciaMax': distancia_max,
                                        'Transportadora': self._parse_transporter_name(file_name),
                                        'Veiculo': str(vehicle), 
                                        'Viagem': viagem_code,
                                        'Tarifa': tarifa, 'Chave': 'N/A & N/A'
                                    })
                            except Exception as inner_e:
                                print(f"Error processing a vehicle column in {file_name}: {inner_e}")
                                continue
                        
                    if processed_rows:
                        all_melted_dfs.append(pd.DataFrame(processed_rows))

                except Exception as e:
                    print(f"Error processing file {file_path} for 'MILK RUN': {e}")
                    continue

        elif 'FAIXA' in fluxo_name:  # Catch '02. FAIXA', 'FAIXA' etc.
            for file_name in os.listdir(fluxo_path):
                if not file_name.lower().endswith(('.xlsx', '.xls')):
                    continue
                file_path = os.path.join(fluxo_path, file_name)
                try:
                    full_df = pd.read_excel(file_path, header=None, engine='openpyxl', keep_default_na=False, na_values=[''])

                    vehicles = full_df.iloc[0].ffill()
                    headers = full_df.iloc[1]

                    # Start reading from row 4 (index 3)
                    data_df = full_df.iloc[3:].copy()
                    data_df.columns = headers

                    rename_map = {}
                    for col in data_df.columns:
                        col_str = str(col).lower().strip()
                        if col_str == 'origem':
                            rename_map[col] = 'Origem'
                        elif col_str == 'destino':
                            rename_map[col] = 'Destino'
                    data_df.rename(columns=rename_map, inplace=True)

                    origem_indices = [i for i, col in enumerate(data_df.columns) if col == 'Origem']
                    destino_indices = [i for i, col in enumerate(data_df.columns) if col == 'Destino']

                    if not origem_indices or not destino_indices:
                        print(f"Skipping file {file_name} due to missing 'Origem' or 'Destino' columns.")
                        continue

                    origem_idx = origem_indices[0]
                    destino_idx = destino_indices[0]

                    processed_rows = []
                    for _, data_row in data_df.iterrows():
                        origem = data_row.iloc[origem_idx]
                        destino_full = data_row.iloc[destino_idx]

                        destino = destino_full
                        distancia_min = None
                        distancia_max = None

                        if isinstance(destino_full, str):
                            # Extract prefix (UF code like "SP", "MG", etc.)
                            destino = destino_full[:2].strip()

                            # Extract "de X a Y" using regex
                            match = re.search(r'de\s*(\d+)\s*a\s*(\d+)', destino_full, flags=re.IGNORECASE)
                            if match:
                                try:
                                    distancia_min = int(match.group(1))
                                    distancia_max = int(match.group(2))
                                except Exception as e:
                                    print(f"Erro ao converter Distancia em '{destino_full}': {e}")

                        for col_idx, _ in enumerate(data_df.columns):
                            if col_idx == origem_idx or col_idx == destino_idx:
                                continue

                            vehicle = vehicles.iloc[col_idx]
                            viagem = headers.iloc[col_idx]
                            tarifa = data_row.iloc[col_idx]

                            if pd.notna(tarifa) and str(tarifa).strip() != "" and pd.notna(vehicle) and pd.notna(viagem):
                                processed_rows.append({
                                    'Nomeacao': 'N/A',
                                    'Fornecedor': 'N/A',
                                    'Origem': origem,
                                    'LocalColeta': 'N/A',
                                    'Destino': destino,
                                    'DistanciaMin': distancia_min,
                                    'DistanciaMax': distancia_max,
                                    'Transportadora': self._parse_transporter_name(file_name),
                                    'Veiculo': str(vehicle),
                                    'Viagem': str(viagem),
                                    'Tarifa': tarifa,
                                    'Chave': str(origem) + ' & ' + str(destino)
                                })

                    if processed_rows:
                        all_melted_dfs.append(pd.DataFrame(processed_rows))

                except Exception as e:
                    print(f"Error processing file {file_path} for 'FAIXA': {e}")
                    continue

        # !------------------------------FLUXO SPOTS -----------------------------------------!
        elif 'SPOTS' in fluxo_name:
            for file_name in os.listdir(fluxo_path):
                if not file_name.lower().endswith(('.xlsx', '.xls')) or file_name.startswith('~$'):
                    continue
                file_path = os.path.join(fluxo_path, file_name)
                try:
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    sheet = wb.active
                    motorista_cols = {}
                    last_vehicle = None
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_val = sheet.cell(row=1, column=col_idx).value
                        if cell_val and str(cell_val).strip():
                            vehicle_name = str(cell_val).strip()
                            if vehicle_name == '0.75':
                                last_vehicle = '3/4'
                            else:
                                last_vehicle = vehicle_name
                        
                        if last_vehicle:
                            # **[FIX]** More robustly parse the motorista value to ensure it's a number.
                            motorista_val_raw = sheet.cell(row=3, column=col_idx).value
                            if motorista_val_raw is not None:
                                try:
                                    # Clean and convert the value to an integer
                                    motorista_clean = int(float(str(motorista_val_raw).strip()))
                                    motorista_cols[col_idx] = (last_vehicle, motorista_clean)
                                except (ValueError, TypeError):
                                    # If conversion fails, it's not a valid motorista column, so skip it
                                    continue

                    header_map = {}
                    data_start_row, header_found_row = 1, -1
                    for r in range(1, min(10, sheet.max_row + 1)):
                        if header_found_row != -1 and r > header_found_row: break
                        for c in range(1, min(20, sheet.max_column + 1)):
                            cell_val = str(sheet.cell(row=r, column=c).value or '').strip().lower()
                            if 'origem' in cell_val: header_map['Origem'] = c
                            elif 'destino' in cell_val: header_map['Destino'] = c
                        if 'Origem' in header_map or 'Destino' in header_map:
                            header_found_row, data_start_row = r, r + 1
                    
                    if 'Origem' not in header_map or 'Destino' not in header_map: continue
                    
                    processed_rows = []
                    for row_idx in range(data_start_row, sheet.max_row + 1):
                        origem = sheet.cell(row=row_idx, column=header_map['Origem']).value
                        destino = sheet.cell(row=row_idx, column=header_map['Destino']).value
                        if not origem or not destino: continue
                        for col_idx, (vehicle, motorista) in motorista_cols.items():
                            tarifa = sheet.cell(row=row_idx, column=col_idx).value
                            if tarifa is not None and str(tarifa).strip() != "":
                                try:
                                    processed_rows.append({
                                        'Transportadora': self._parse_transporter_name(file_name),
                                        'Veiculo': vehicle, 'Motorista': motorista,
                                        'Origem': str(origem).strip(), 'Destino': str(destino).strip(),
                                        'Distancia': float(motorista),
                                        'Tarifa': float(tarifa),
                                        'Nomeacao': 'N/A', 'Fornecedor': 'N/A', 'LocalColeta': 'N/A', 'Viagem': 'N/A',
                                        'Chave': f"{origem} & {destino}"
                                    })
                                except (ValueError, TypeError): continue
                    if processed_rows: all_melted_dfs.append(pd.DataFrame(processed_rows))
                except Exception as e:
                    print(f"Error processing file {file_path} for 'SPOTS': {e}")
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        else:

            # --- ORIGINAL Logic for '01', '03', etc. ---
            geoship_df = None
            try:
                parent_folder = os.path.dirname(self.base_folder)

                # Look for Geoship table in that parent folder
                geoship_filename = next(
                    (f for f in os.listdir(parent_folder)
                    if 'geoshiptable' in f.lower() and f.endswith(('.xlsx', '.xls'))),
                    None
                )

                if geoship_filename:
                    geoship_full_path = os.path.join(parent_folder, geoship_filename)
                    geoship_df = pd.read_excel(geoship_full_path, engine='openpyxl')
                    geoship_df = geoship_df.rename(columns={
                        'Fornecedor': 'Fornecedor_geoship',
                        'Km Total': 'Distancia_geoship',
                        'Destino Materiais': 'Destino_geoship'
                    })
                    print(f"✅ Successfully loaded data source: '{geoship_filename}'.")
                else:
                    print("⚠️ Warning: 'GeoshipTable' file not found.")

            except FileNotFoundError:
                print(f" Error: Directory not found at '{parent_folder}'.")
            except Exception as e:
                print(f" Error loading GeoshipTable: {e}.")


            # --- ORIGINAL Logic for '01', '03', etc. ---
            for file_name in os.listdir(fluxo_path):
                if not file_name.lower().endswith(('.xlsx', '.xls')) or file_name.startswith('~'):
                    continue
                file_path = os.path.join(fluxo_path, file_name)
                try:
                    header_df = pd.read_excel(file_path, header=None, nrows=2, engine='openpyxl')
                    header_df.iloc[0] = header_df.iloc[0].ffill()
                    new_columns = []
                    for i in range(len(header_df.columns)):
                        top_header = str(header_df.iloc[0, i]).upper().strip()
                        bottom_header = str(header_df.iloc[1, i]).upper().strip()
                        if 'UNNAMED' in top_header or top_header == 'NAN':
                            new_columns.append(bottom_header.lower())
                        elif 'UNNAMED' in bottom_header or bottom_header == 'NAN':
                            new_columns.append(top_header)
                        else:
                            new_columns.append(f"{top_header}_{bottom_header}")

                    df = pd.read_excel(file_path, header=None, skiprows=2, engine='openpyxl')
                    min_cols = min(len(df.columns), len(new_columns))
                    df = df.iloc[:, :min_cols]
                    df.columns = new_columns[:min_cols]

                    df.columns = [col.strip() for col in df.columns]

                    tipo_fluxo_col = next((col for col in df.columns if 'tipo de fluxo' in col.lower()), None)

                    fornecedor_col_name = next((c for c in df.columns if 'fornecedor' in c.lower() and 'codigo' not in c.lower()), None)

                    id_cols_map = {
                        'Nomeacao': next((c for c in df.columns if ('nomeação' in c.lower()) or ('nomeacao' in c.lower())), 'Nomeacao'),
                        'Origem': next((c for c in df.columns if ('cidade de coleta' in c.lower()) or ('cidade_coleta' in c.lower())), 'Origem'),
                        'LocalColeta': next((c for c in df.columns if ('local de coleta' in c.lower()) or ('local_coleta' in c.lower())), 'LocalColeta'),
                        'Destino': next((c for c in df.columns if 'destino materiais' in c.lower()), 'Destino'),
                        'Distancia': next((c for c in df.columns if 'distância' in c.lower()), 'Distancia'),
                    }
                    if fornecedor_col_name:
                        id_cols_map['Fornecedor'] = fornecedor_col_name

                    df.rename(columns={v: k for k, v in id_cols_map.items() if v in df.columns}, inplace=True)

                    if 'Fornecedor' not in df.columns:
                        df['Fornecedor'] = 'N/A'

                    id_vars = list(id_cols_map.keys())
                    if tipo_fluxo_col:
                        id_vars.append(tipo_fluxo_col)

                    df['Transportadora'] = self._parse_transporter_name(file_name)
                    id_vars.append('Transportadora')

                    value_vars = [col for col in df.columns if '_' in col and col not in id_vars]

                    melted_df = df.melt(
                        id_vars=[v for v in id_vars if v in df.columns],
                        value_vars=value_vars,
                        var_name='Veiculo_Viagem',
                        value_name='Tarifa'
                    )

                    melted_df[['Veiculo', 'Viagem']] = melted_df['Veiculo_Viagem'].str.split('_', expand=True, n=1)
                    melted_df.drop('Veiculo_Viagem', axis=1, inplace=True)
                    melted_df['Chave'] = melted_df['Origem'].astype(str) + ' & ' + melted_df['Destino'].astype(str)

                    # Replace Geoship Rows
                    # Replace Geoship Rows
                    if tipo_fluxo_col and tipo_fluxo_col in melted_df.columns and geoship_df is not None:
                        is_geoship = melted_df[tipo_fluxo_col].astype(str).str.lower().str.contains('geoship', na=False)
                        geoship_matches = melted_df[is_geoship]

                        non_geoship = melted_df[~is_geoship]
                        updated_rows = []

                        # For each row where Tipo de Fluxo contains Geoship
                        for _, row in geoship_matches.iterrows():
                            tipo_fluxo_value = str(row[tipo_fluxo_col]).strip()
                            # Search Geoship table for rows where Tipo de Fluxo matches (or contains) the tipo_fluxo_value
                            # Here we assume geoship_df has a column like 'Tipo de Fluxo' or similar to match
                            # Adjust the column name below as needed, for example 'Tipo de Fluxo' or 'Tipo_de_Fluxo' or something else
                            # We'll assume it's 'Tipo de Fluxo' exactly for now
                            geoship_key_col = next((col for col in geoship_df.columns if 'tipo' in col.lower() and 'fluxo' in col.lower()), None)
                            if geoship_key_col is None:
                                # fallback: just match on some column named 'Geoship' if exists
                                geoship_key_col = next((col for col in geoship_df.columns if 'geoship' in col.lower()), None)
                            
                            if geoship_key_col:
                                # Find rows in geoship_df that contain the tipo_fluxo_value (case insensitive)
                                matched_geo_rows = geoship_df[
                                            geoship_df[geoship_key_col].astype(str).str.lower() == tipo_fluxo_value.lower()
                                        ]

                            else:
                                matched_geo_rows = pd.DataFrame()  # no matches possible

                            if matched_geo_rows.empty:
                                # No match found, keep original row as is
                                updated_rows.append(row)
                            else:
                                # For each matching geoship row, create a new combined row
                                for _, geo_row in matched_geo_rows.iterrows():
                                    new_row = row.copy()
                                    # Update columns with geoship data, rename accordingly if needed
                                    # Match columns you renamed earlier: Fornecedor_geoship, Distancia_geoship, Destino_geoship, plus others if relevant
                                    # You can add any other columns you want to override from geo_row here
                                    new_row['Fornecedor'] = geo_row.get('Fornecedor_geoship', new_row.get('Fornecedor', 'N/A'))
                                    new_row['Distancia'] = geo_row.get('Distancia_geoship', new_row.get('Distancia', None))
                                    new_row['Origem'] = geo_row.get('CNPJ Origem', new_row.get('Origem', None))
                                    new_row['Destino'] = geo_row.get('Destino_geoship', new_row.get('Destino', None))
                                    # You can add other columns from geo_row if needed

                                    updated_rows.append(new_row)

                        # Combine all non-geoship rows and expanded geoship rows
                        melted_df = pd.concat([non_geoship, pd.DataFrame(updated_rows)], ignore_index=True).drop(columns=[tipo_fluxo_col], errors='ignore')



                    all_melted_dfs.append(melted_df)

                except Exception as e:
                    print(f"Error processing file {file_path}: {e}")
                    continue


        
        if not all_melted_dfs:
            return {'success': False, 'message': 'Nenhum arquivo de tarifa válido foi encontrado.'}

        master_df = pd.concat(all_melted_dfs, ignore_index=True).dropna(subset=['Tarifa'])
        
        # --- NEW & IMPROVED DATA CLEANING ---
        str_cols = ['Nomeacao', 'Fornecedor', 'Origem', 'Local Coleta', 'Destino', 'Veiculo', 'Viagem']
        for col in str_cols:
            if col in master_df.columns:
                master_df[col] = master_df[col].astype(str).str.strip()

        if 'Viagem' in master_df.columns:
            master_df['Viagem'] = master_df['Viagem'].str.upper()

        special_cols = ['Nomeacao', 'Fornecedor']
        for col in special_cols:
             if col in master_df.columns:
                master_df[col] = master_df[col].str.upper().str.replace(' ', '').fillna('N/A')

        if 'Nomeacao' in master_df.columns:
            master_df['Nomeacao'] = master_df['Nomeacao'].replace('PRINCIPALCARRETA', 'PRINCIPAL')
        # --- END OF CLEANING ---

        for col in ['Tarifa', 'Distancia']:
            if col in master_df.columns:
                master_df[col] = pd.to_numeric(master_df[col], errors='coerce')

        filters = {
            'Nomeacao': sorted(master_df['Nomeacao'].dropna().unique().tolist()),
            'Fornecedores': sorted(master_df['Fornecedor'].dropna().unique().tolist()),
            'Origem': sorted(master_df['Origem'].dropna().unique().tolist()),
            'LocaisColeta': sorted(master_df['LocalColeta'].dropna().unique().tolist()),
            'Destino': sorted(master_df['Destino'].dropna().unique().tolist()),
            'Veiculos': sorted(master_df['Veiculo'].dropna().unique().tolist())
        }

        # **[FIX]** More robustly create the Motoristas filter list
        if 'Motorista' in master_df.columns:
            try:
                # Ensure values are clean integers before creating the unique list
                unique_motoristas = master_df['Motorista'].dropna().astype(int).unique()
                filters['Motoristas'] = sorted(unique_motoristas.tolist())
            except (ValueError, TypeError):
                # If conversion fails for any reason, return an empty list
                filters['Motoristas'] = []
        
        response = {'success': True, 'filters': filters}
        self.fluxo_data[fluxo_name] = {'df': master_df, 'filters_response': response}
        
        return response

   

    def calculate_tariffs(self, params):
        fluxo_name = params['fluxo']
        if not fluxo_name or fluxo_name not in self.fluxo_data: return []
        df = self.fluxo_data[fluxo_name]['df'].copy()
        
        # --- Apply Filters ---
        if params.get('fornecedor'): df = df[df['Fornecedor'] == params['fornecedor']]
        if params.get('nomeacao'): df = df[df['Nomeacao'] == params['nomeacao']]
        if params.get('origem'): df = df[df['Origem'] == params['origem']]
        if params.get('local_coleta') and 'LocalColeta' in df.columns: df = df[df['LocalColeta'] == params['local_coleta']]
        if params.get('destino'): df = df[df['Destino'] == params['destino']]
        if params.get('veiculo'): df = df[df['Veiculo'] == params['veiculo']]
        
        if 'SPOTS' not in fluxo_name.upper():
            if params.get('calc_type'): 
                df = df[df['Viagem'] == params['calc_type']]

        if params.get('motorista') and 'Motorista' in df.columns:
            try: df = df[df['Motorista'] == int(params['motorista'])]
            except (ValueError, TypeError): pass

        if df.empty: return []

        # --- Perform Calculations ---
        try:
            new_distance = float(params.get('km_value'))
            if new_distance <= 0: new_distance = None
        except (ValueError, TypeError): new_distance = None

        if new_distance:
            df['Tarifa Real'] = pd.NA
            if ('FAIXA' in fluxo_name.upper() or 'MILK RUN' in fluxo_name.upper()) and 'DistanciaMin' in df.columns:
                mask = (df['DistanciaMin'] <= new_distance) & (df['DistanciaMax'] >= new_distance)
                if 'MILK RUN' in fluxo_name.upper():
                    df.loc[mask, 'Tarifa Real'] = new_distance * df.loc[mask, 'Tarifa']
                else:
                    df.loc[mask, 'Tarifa Real'] = df.loc[mask, 'Tarifa']
                df = df[mask].copy()
            elif 'Distancia' in df.columns:
                mask = (df['Distancia'].notna()) & (df['Distancia'] > 0)
                df.loc[mask, 'Tarifa Real'] = (new_distance * df.loc[mask, 'Tarifa']) / df.loc[mask, 'Distancia']
                df.loc[mask, 'Distancia'] = new_distance
        else:
            if 'SPOTS' in fluxo_name.upper() and 'Distancia' in df.columns:
                df['Distancia'] = pd.NA

        # <<< START OF CORRECTION
        # Filter out any results where the calculated tariff is 0
        if 'Tarifa Real' in df.columns:
            df = df[df['Tarifa Real'] > 0]
        # <<< END OF CORRECTION

        # --- [NEW] Unified Final Processing, Sorting, and Column Arrangement ---
        
        # 1. Determine the primary column for sorting
        sort_col = 'Tarifa Real' if 'Tarifa Real' in df.columns and df['Tarifa Real'].notna().any() else 'Tarifa'

        # 2. Sort the entire result set by the best tariff, lowest first
        df_sorted = df.sort_values(by=sort_col, ascending=True)

        # 3. Define the exact columns and order for the final display
        final_display_cols = [
            'Origem',
            'Destino',
            'Veiculo',
            'Motorista',
            'Transportadora',
            'Distancia',
            'Tarifa',
            'Tarifa Real'
        ]

        # 4. Select only the columns that exist in the current dataframe
        existing_cols = [col for col in final_display_cols if col in df_sorted.columns]
        self.last_results_df = df_sorted[existing_cols].copy()

        # 5. Round the tariff columns just before returning
        if 'Tarifa Real' in self.last_results_df.columns:
            self.last_results_df['Tarifa Real'] = pd.to_numeric(self.last_results_df['Tarifa Real'], errors='coerce').round(2)
        if 'Tarifa' in self.last_results_df.columns:
            self.last_results_df['Tarifa'] = pd.to_numeric(self.last_results_df['Tarifa'], errors='coerce').round(2)
        
        return self.last_results_df.to_dict('records')

    def export_to_excel(self):
        if self.last_results_df is None or self.last_results_df.empty:
            return {'success': False, 'message': 'Não há dados para exportar.'}
        if not self._window:
            return {'success': False, 'message': 'Erro: Referência da janela não encontrada.'}
        try:
            file_path_tuple = self._window.create_file_dialog(webview.SAVE_DIALOG, directory=os.path.expanduser('~'), save_filename='cotacao_tarifas.xlsx')
            if not file_path_tuple:
                return {'success': False, 'message': 'Exportação cancelada.'}
            user_choice = file_path_tuple[0]
            filename = os.path.basename(user_choice)
            if not filename or ":" in filename:
                downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
                os.makedirs(downloads_path, exist_ok=True)
                final_path = os.path.join(downloads_path, 'cotacao_tarifas.xlsx')
            else:
                final_path = user_choice
                if not final_path.lower().endswith('.xlsx'): final_path += '.xlsx'
            self.last_results_df.to_excel(final_path, index=False, engine='openpyxl')
            return {'success': True, 'message': f'Sucesso! Arquivo salvo em: {final_path}'}
        except Exception as e:
            return {'success': False, 'message': f'Erro ao exportar: {e}'}

if __name__ == '__main__':
    api = Api()

    window = webview.create_window(
        'Calculador de Tarifas Unificado',
        html=HTML_CONTENT,
        js_api=api,
        width=1400,
        height=900,
        resizable=True
    )
    
    api._window = window

    webview.start()